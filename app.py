import io, re, unicodedata
from datetime import datetime
from pathlib import Path
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Indicador Forecast x Realizado", page_icon="📦", layout="wide")
MONTHS={"JAN":1,"FEV":2,"MAR":3,"ABR":4,"MAI":5,"JUN":6,"JUL":7,"AGO":8,"SET":9,"OUT":10,"NOV":11,"DEZ":12}
MONTH_NAMES_PT=["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
FILTERS=[("Considerar volumetria","Considerar no Forecast"),("CD_Ajustado","CD"),("Empresa_Ajustado","Empresa"),("Canal de Atuação (Novo)","Canal de Atuação (Novo)"),("Canal Consolidador (Novo)","Agrupamento de Canal"),("Drive de cobrança","Drive de cobrança"),("Diretoria","Diretoria"),("OPL_Ajustado","Fornecedor"),("UF","UF"),("Movimento","Movimento"),("Área","Área"),("Resp. Fcst Planejamento","Responsável FCST")]
SERIES={"real":"Realizado","fcst":"FCST","opl":"FCST Envio OPL"}; COLORS={"real":"#1f36ad","fcst":"#2386dd","opl":"#ff8f00"}
st.markdown('''<style>
.stApp{background:#fff}.block-container{max-width:1600px;padding-top:1rem}.header{background:#fff;padding:22px 28px 18px;border-radius:0 0 16px 16px;margin-bottom:15px;box-shadow:0 2px 9px #00000010}.header h1{color:#30323d;font-size:2.45rem;line-height:1.15;margin:0;font-weight:800}.header p{color:#8a8e99;font-size:.88rem;margin:18px 0 0}[data-testid="stMetric"]{background:#fff;border:1px solid #d8dee7;border-radius:15px;padding:12px 16px;box-shadow:0 3px 10px #2230440c;min-height:125px}[data-testid="stMetricLabel"]{font-size:.82rem;min-height:2.65rem;align-items:flex-start}[data-testid="stMetricLabel"] p{white-space:normal!important;overflow:visible!important;text-overflow:clip!important;line-height:1.2!important}[data-testid="stMetricValue"]{font-size:2rem}[data-testid="stSidebar"]{background:#eef0f4;border-right:1px solid #dde1e7}[data-testid="stSidebar"]>div:first-child{background:#eef0f4}[data-testid="stSidebar"] [data-baseweb="select"]>div,[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"],[data-testid="stSidebar"] [data-testid="stDateInput"]>div>div{background:#fff;border-radius:10px}div[data-testid="stDataFrame"],div[data-testid="stPlotlyChart"]{background:#fff;border:1px solid #dce2e8;border-radius:15px;padding:9px}.section{font-size:1.08rem;font-weight:750;margin:.7rem 0 .65rem;color:#263449}.card-caption{font-size:.82rem;color:#707887;margin:-.25rem 0 .65rem}</style>''',unsafe_allow_html=True)

def norm(v): return unicodedata.normalize("NFKD",str(v)).encode("ascii","ignore").decode().strip().casefold()
def parse_month(h):
 m=re.search(r"(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)/(\d{2})$",str(h).strip().upper()); return pd.Timestamp(2000+int(m.group(2)),MONTHS[m.group(1)],1) if m else None
def series_type(h):
 t=str(h).strip().upper()
 if re.match(r"^REAL ",t): return "real"
 if re.match(r"^FCST \(ENVIO OPL\) ",t): return "opl"
 if re.match(r"^FCST ",t) and not t.endswith("(ANT)"): return "fcst"
 return None
@st.cache_data(show_spinner="Lendo e preparando a base Farol...")
def load_farol(raw):
 wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True); ws=wb["Farol"]
 headers=[ws.cell(4,c).value for c in range(1,ws.max_column+1)]; wide=pd.DataFrame(list(ws.iter_rows(min_row=5,values_only=True)),columns=headers); wide=wide.loc[:,wide.columns.notna()].dropna(how="all")
 meta={c:(series_type(c),parse_month(c)) for c in wide.columns if series_type(c) and parse_month(c) is not None}; dims=[c for c,_ in FILTERS if c in wide.columns]
 long=wide[dims+list(meta)].melt(id_vars=dims,var_name="coluna",value_name="valor"); long["serie"]=long["coluna"].map(lambda c:meta[c][0]); long["mes"]=long["coluna"].map(lambda c:meta[c][1]); long["valor"]=pd.to_numeric(long["valor"],errors="coerce").fillna(0)
 for c in dims: long[c]=long[c].fillna("Não informado").astype(str).str.strip()
 return long.drop(columns="coluna"),len(wide),len(meta)
def num(v): return f"{v:,.0f}".replace(",",".")
def pct(v): return "-" if pd.isna(v) else f"{v:.2%}".replace(".",",")
def accuracy(r,x): return np.nan if x==0 else max(0,1-abs(r-x)/abs(x))
def month_label(v): return pd.Timestamp(v).strftime("%m/%Y")
def month_short(v): d=pd.Timestamp(v); return f"{MONTH_NAMES_PT[d.month-1]}/{d:%y}"
def month_long(v): d=pd.Timestamp(v); return f"{MONTH_NAMES_PT[d.month-1]}/{d:%Y}"
def aggregate(df):
 o=df.groupby(["mes","serie"],as_index=False)["valor"].sum().pivot(index="mes",columns="serie",values="valor").fillna(0).reset_index()
 for c in SERIES:
  if c not in o:o[c]=0.
 return o.sort_values("mes")
def positions(data,keys):
 out={k:[] for k in keys}
 for i in range(len(data)):
  present=sorted([(k,data.iloc[i][k]) for k in keys if data.iloc[i][k]!=0],key=lambda x:x[1]); assigned={}
  for j,(k,_) in enumerate(present): assigned[k]="bottom center" if j==0 and len(present)>1 else ("top center" if j==len(present)-1 else ("middle right" if i%2==0 else "middle left"))
  for k in keys: out[k].append(assigned.get(k,"top center"))
 return out
def single_positions(values):
 vals=list(values); out=[]
 for i,v in enumerate(vals):
  prev=vals[i-1] if i>0 and not pd.isna(vals[i-1]) else None; nxt=vals[i+1] if i<len(vals)-1 and not pd.isna(vals[i+1]) else None
  out.append("bottom center" if prev is not None and nxt is not None and v<=prev and v<=nxt else "top center")
 return out
def render_cards(title,subtitle,m,comparison):
 valid=m[m.real>0]; t=valid[["real","fcst","opl"]].sum() if len(valid) else pd.Series({"real":0.,"fcst":0.,"opl":0.}); cards=[("Realizado",num(t.real),None)]
 if comparison in ("FCST e Realizado","FCST, FCST Envio OPL e Realizado"):
  d=t.real-t.fcst; cards += [("FCST",num(t.fcst),None),("Desvio FCST",num(d),pct(d/t.fcst) if t.fcst else None),("Acuracidade FCST",pct(accuracy(t.real,t.fcst)),None)]
 if comparison in ("FCST Envio OPL e Realizado","FCST, FCST Envio OPL e Realizado"):
  d=t.real-t.opl; cards += [("FCST Envio OPL",num(t.opl),None),("Desvio FCST Envio OPL",num(d),pct(d/t.opl) if t.opl else None),("Acuracidade FCST Envio OPL",pct(accuracy(t.real,t.opl)),None)]
 st.markdown(f'<div class="section">{title}</div><div class="card-caption">{subtitle}</div>',unsafe_allow_html=True)
 for col,(a,b,c) in zip(st.columns(len(cards)),cards): col.metric(a,b,c)

path=Path("Farol.xlsx")
if path.exists():
 raw=path.read_bytes(); updated=datetime.fromtimestamp(path.stat().st_mtime)
else:
 st.error("Base não encontrada. Coloque o arquivo 'Farol.xlsx' na raiz do repositório."); st.stop()
base,nrows,ncols=load_farol(raw); st.markdown(f'<div class="header"><h1>📦 Indicador Forecast x Realizado</h1></div>',unsafe_allow_html=True)
st.sidebar.markdown("### Filtros"); activity=base.groupby("mes").valor.apply(lambda x:x.abs().sum()); active=activity[activity>0].index.sort_values(); real=base[base.serie=="real"].groupby("mes").valor.sum(); received=real[real>0].index.sort_values(); last=received.max() if len(received) else active.max(); opts=list(active)[::-1]
period=st.sidebar.selectbox("Período da evolução",["Últimos 3 meses","Últimos 6 meses","Últimos 9 meses","Últimos 12 meses","Livre escolha"],index=3); selected_month=st.sidebar.selectbox("Mês/Ano para os cards",opts,index=opts.index(last),format_func=month_label); comparison=st.sidebar.selectbox("Comparação",["FCST, FCST Envio OPL e Realizado","FCST e Realizado","FCST Envio OPL e Realizado"])
filtered=base
for c,label in FILTERS:
 if c in filtered:
  options=sorted(filtered[c].dropna().unique(),key=norm); default=["Sim"] if c=="Considerar volumetria" and "Sim" in options else (["OUT"] if c=="Movimento" and "OUT" in options else []); chosen=st.sidebar.multiselect(label,options,default=default,placeholder="Todos")
  if chosen: filtered=filtered[filtered[c].isin(chosen)]
active=filtered.groupby("mes").valor.apply(lambda x:x.abs().sum()); active=active[active>0].index.sort_values()
if period=="Livre escolha":
 a,b=st.sidebar.columns(2); start=a.date_input("Início",active.min().date()); end=b.date_input("Fim",active.max().date()); evolution=filtered[filtered.mes.between(pd.Timestamp(start),pd.Timestamp(end))]
else:
 n=int(re.search(r"\d+",period).group()); end=active.max(); start=end-pd.DateOffset(months=n-1); evolution=filtered[filtered.mes.between(start,end)]
monthly=aggregate(evolution); month_data=aggregate(filtered[filtered.mes==selected_month]); ref="opl" if comparison=="FCST Envio OPL e Realizado" else "fcst"; keys=["fcst","real"] if comparison=="FCST e Realizado" else (["opl","real"] if comparison=="FCST Envio OPL e Realizado" else ["fcst","opl","real"])
render_cards("Indicadores do mês selecionado",month_label(selected_month),month_data,comparison); render_cards("Indicadores do período de evolução",f"{month_label(monthly.mes.min())} a {month_label(monthly.mes.max())}",monthly,comparison)

def line_chart(title,series_data,y_format,colors,percent=False):
 st.markdown(f'<div class="section">{title}</div>',unsafe_allow_html=True); fig=go.Figure(); pos=positions(series_data,list(colors))
 for key,color in colors.items():
  y=series_data[key].where(series_data[key]!=0,np.nan); labels=["" if pd.isna(v) else (pct(v) if percent else num(v)) for v in y]
  fig.add_trace(go.Scatter(x=series_data.mes,y=y,mode="lines+markers+text",name=key,line=dict(color=color,width=3),marker=dict(size=7),text=labels,textposition=pos[key],textfont=dict(size=18,color=color),cliponaxis=False,connectgaps=False))
 fig.update_layout(height=550,margin=dict(l=55,r=55,t=90,b=55),hovermode="x unified",legend=dict(orientation="h",y=1.12,font=dict(size=14)),font=dict(size=14,color="#263449"),paper_bgcolor="white",plot_bgcolor="white")
 fig.update_xaxes(tickmode="array",tickvals=series_data.mes,ticktext=[month_short(v) for v in series_data.mes],gridcolor="#eef1f4",tickfont=dict(size=13)); fig.update_yaxes(gridcolor="#e7ebf0",tickformat=y_format,tickfont=dict(size=13))
 st.plotly_chart(fig,use_container_width=True)

line_chart("Evolução mensal",monthly[["mes"]+keys].rename(columns={k:SERIES[k] for k in keys}),",.0f",{SERIES[k]:COLORS[k] for k in keys})
# Novo gráfico solicitado: desvio percentual imediatamente abaixo da evolução mensal.
dev=pd.DataFrame({"mes":monthly.mes})
dev_colors={}
if comparison in ("FCST e Realizado","FCST, FCST Envio OPL e Realizado"):
 dev["Desvio FCST"] = np.where((monthly.real>0)&(monthly.fcst>0),(monthly.real-monthly.fcst)/monthly.fcst,np.nan); dev_colors["Desvio FCST"] = COLORS["fcst"]
if comparison in ("FCST Envio OPL e Realizado","FCST, FCST Envio OPL e Realizado"):
 dev["Desvio FCST Envio OPL"] = np.where((monthly.real>0)&(monthly.opl>0),(monthly.real-monthly.opl)/monthly.opl,np.nan); dev_colors["Desvio FCST Envio OPL"] = COLORS["opl"]
line_chart("Evolução do desvio",dev,".0%",dev_colors,percent=True)

acc=pd.DataFrame({"mes":monthly.mes}); acc_colors={}
if "fcst" in keys: acc["Acuracidade FCST"]=np.where((monthly.real>0)&(monthly.fcst>0),np.maximum(0,1-(monthly.real-monthly.fcst).abs()/monthly.fcst.abs()),np.nan); acc_colors["Acuracidade FCST"]=COLORS["fcst"]
if "opl" in keys: acc["Acuracidade FCST Envio OPL"]=np.where((monthly.real>0)&(monthly.opl>0),np.maximum(0,1-(monthly.real-monthly.opl).abs()/monthly.opl.abs()),np.nan); acc_colors["Acuracidade FCST Envio OPL"]=COLORS["opl"]
line_chart("Evolução da acuracidade",acc,".0%",acc_colors,percent=True)
st.info("**Desvio FCST:** diferença entre o volume Realizado e o volume previsto no FCST. | **Acuracidade FCST:** mede o quanto o FCST ficou próximo do Realizado; quanto mais perto de 100%, maior a precisão.")

summary=monthly.copy(); summary["Desvio"]=summary.real-summary[ref]; summary["Desvio %"]=np.where(summary[ref]!=0,summary["Desvio"]/summary[ref],np.nan); summary["Acuracidade"]=[accuracy(r,x) for r,x in zip(summary.real,summary[ref])]; summary=summary.rename(columns={"mes":"Mês","real":"Realizado","fcst":"FCST","opl":"FCST Envio OPL"}); show=["Mês"]+[SERIES[k] for k in keys]+["Desvio","Desvio %","Acuracidade"]
left,right=st.columns([1.6,1])
with left:
 st.markdown('<div class="section">Resumo mensal</div>',unsafe_allow_html=True); disp=summary[show].copy(); disp["Mês"]=disp["Mês"].map(month_long)
 for c in ["FCST","FCST Envio OPL","Realizado","Desvio"]:
  if c in disp: disp[c]=disp[c].map(num)
 for c in ["Desvio %","Acuracidade"]: disp[c]=disp[c].map(pct)
 st.dataframe(disp,hide_index=True,use_container_width=True)
with right:
 st.markdown('<div class="section">Maiores desvios por CD</div>',unsafe_allow_html=True); cd="CD_Ajustado"; rank=evolution.groupby([cd,"serie"],as_index=False).valor.sum().pivot(index=cd,columns="serie",values="valor").fillna(0).reset_index(); rank=rank[rank.real>0]; rank["Desvio %"]=np.where(rank[ref]!=0,(rank.real-rank[ref])/rank[ref],np.nan); rank=rank.assign(ordem=lambda x:x["Desvio %"].abs()).sort_values("ordem",ascending=False).head(12).rename(columns={cd:"CD","real":"Realizado",ref:SERIES[ref]}); rd=rank[["CD","Realizado",SERIES[ref],"Desvio %"]].copy(); rd["Realizado"]=rd["Realizado"].map(num); rd[SERIES[ref]]=rd[SERIES[ref]].map(num); rd["Desvio %"]=rd["Desvio %"].map(pct); st.dataframe(rd,hide_index=True,use_container_width=True)
out=io.BytesIO()
with pd.ExcelWriter(out,engine="openpyxl") as w: summary[show].to_excel(w,index=False,sheet_name="Resumo mensal"); evolution.to_excel(w,index=False,sheet_name="Base filtrada")
st.download_button("Baixar resultado em Excel",out.getvalue(),"Resumo_Farol.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
